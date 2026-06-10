from datetime import date

import polars as pl
import pytest
from pandera.errors import SchemaError
from polars.testing import assert_schema_equal

from moneypype.etl_excel import run, _validate_input
from moneypype.schemas import TRANSACTIONS_SCHEMA


@pytest.fixture
def categories_map_file(tmp_path):
    content = "Category,Type\nSalary,Income\nGroceries,Needs\nShopping,Wants\n"
    path = tmp_path / "categories_map.csv"
    path.write_text(content)
    return str(path)


@pytest.fixture
def xlsx_file(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()

    ws_income = wb.active
    ws_income.title = "Dochody"
    ws_income.append(["Dochody"])
    ws_income.append([
        "Data i godzina", "Konto", "Kategoria",
        "Kwota w walucie domyślnej", "Waluta domyślna",
        "Kwota w walucie konta", "Waluta konta",
        "Komentarz", "Etykietki",
    ])
    ws_income.append([
        date(2026, 1, 15), "Savings", "Salary",
        1000.0, "PLN", 1000.0, "PLN", "Salary", None,
    ])

    ws_expenses = wb.create_sheet("Wydatki")
    ws_expenses.append(["Wydatki"])
    ws_expenses.append([
        "Data i godzina", "Konto", "Kategoria",
        "Kwota w walucie domyślnej", "Waluta domyślna",
        "Kwota w walucie konta", "Waluta konta",
        "Komentarz", "Etykietki",
    ])
    ws_expenses.append([
        date(2026, 1, 20), "Alior", "Groceries",
        50.0, "PLN", 50.0, "PLN", None, None,
    ])

    ws_transfers = wb.create_sheet("Przelewy")
    ws_transfers.append(["Przelewy"])
    ws_transfers.append([
        "Data i godzina", "Wychodzące", "Kwota w walucie wychodzącej",
        "Waluta wychodząca", "Przychodzące", "Waluta w kwocie przychodzącej",
        "Komentarz",
    ])
    ws_transfers.append([
        date(2026, 1, 25), "Alior", 200.0, "PLN", "Savings", None, "Transfer",
    ])

    path = tmp_path / "test.xlsx"
    wb.save(path)
    return str(path)


def test_run_schema(tmp_path, xlsx_file, categories_map_file):
    dest = str(tmp_path / "out.parquet")
    result = run(xlsx_file, dest, categories_map_file)
    assert_schema_equal(result.schema, TRANSACTIONS_SCHEMA)


def test_run_transfer_expands_to_two_rows(
    tmp_path, xlsx_file, categories_map_file
):
    dest = str(tmp_path / "out.parquet")
    result = run(xlsx_file, dest, categories_map_file)
    transfers = result.filter(pl.col("type") == "Transfer")
    assert len(transfers) == 2


def test_run_expense_amount_is_negative(
    tmp_path, xlsx_file, categories_map_file
):
    dest = str(tmp_path / "out.parquet")
    result = run(xlsx_file, dest, categories_map_file)
    expense = result.filter(pl.col("category") == "Groceries")
    assert expense["amount"][0] < 0


def test_run_income_amount_is_positive(
    tmp_path, xlsx_file, categories_map_file
):
    dest = str(tmp_path / "out.parquet")
    result = run(xlsx_file, dest, categories_map_file)
    income = result.filter(pl.col("category") == "Salary")
    assert income["amount"][0] > 0


def test_run_amounts_in_cents(tmp_path, xlsx_file, categories_map_file):
    dest = str(tmp_path / "out.parquet")
    result = run(xlsx_file, dest, categories_map_file)
    income = result.filter(pl.col("category") == "Salary")
    assert income["amount"][0] == 100000


def test_run_transfer_to_row_has_null_fx_amount(
    tmp_path, xlsx_file, categories_map_file
):
    dest = str(tmp_path / "out.parquet")
    result = run(xlsx_file, dest, categories_map_file)
    # to-row has positive amount; from-row has negative amount
    to_row = result.filter(
        (pl.col("type") == "Transfer") & (pl.col("amount") > 0)
    )
    assert to_row["amount_fx_ccy"][0] is None


def test_run_sorted_by_date(tmp_path, xlsx_file, categories_map_file):
    dest = str(tmp_path / "out.parquet")
    result = run(xlsx_file, dest, categories_map_file)
    dates = result["date"].to_list()
    assert dates == sorted(dates)


def test_input_validation_raises():
    bad_data = pl.DataFrame({
        "date": pl.Series([date(2026, 1, 1)], dtype=pl.Date),
        "account": pl.Series(["Savings"], dtype=pl.String),
        "category": pl.Series(["Salary"], dtype=pl.String),
        "type": pl.Series(["Income"], dtype=pl.String),
        "note": pl.Series([None], dtype=pl.String),
        "currency": pl.Series(["PLN"], dtype=pl.String),
        "amount": pl.Series(["not_a_number"], dtype=pl.String),
        "ref_currency_amount": pl.Series([None], dtype=pl.Float64),
        "label": pl.Series([None], dtype=pl.String),
    })
    with pytest.raises(SchemaError):
        _validate_input(bad_data)
