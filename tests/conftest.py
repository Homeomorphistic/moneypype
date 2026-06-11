from datetime import date
from pathlib import Path

import pytest


@pytest.fixture
def csv_file():
    return Path(__file__).parent / "fixtures" / "sample.csv"


@pytest.fixture
def categories_map_file(tmp_path):
    content = "Category,Type\nSalary,Income\nGroceries,Needs\nShopping,Wants\n"
    path = tmp_path / "categories_map.csv"
    path.write_text(content)
    return str(path)


@pytest.fixture
def xlsx_file(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()

    ws_income = wb.active
    ws_income.title = "Dochody"
    ws_income.append(["Dochody"])
    ws_income.append([
        "Data i godzina", "Konto", "Kategoria",
        "Kwota w walucie domyślnej", "Waluta domyślna",
        "Kwota w walucie konta", "Waluta konta",
        "Komentarz", "Etykietki",
    ])
    ws_income.append([
        date(2026, 1, 15), "Savings", "Salary",
        1000.0, "PLN", 1000.0, "PLN", "Salary", None,
    ])

    ws_expenses = wb.create_sheet("Wydatki")
    ws_expenses.append(["Wydatki"])
    ws_expenses.append([
        "Data i godzina", "Konto", "Kategoria",
        "Kwota w walucie domyślnej", "Waluta domyślna",
        "Kwota w walucie konta", "Waluta konta",
        "Komentarz", "Etykietki",
    ])
    ws_expenses.append([
        date(2026, 1, 20), "Alior", "Groceries",
        50.0, "PLN", 50.0, "PLN", None, None,
    ])

    ws_transfers = wb.create_sheet("Przelewy")
    ws_transfers.append(["Przelewy"])
    ws_transfers.append([
        "Data i godzina", "Wychodzące", "Kwota w walucie wychodzącej",
        "Waluta wychodząca", "Przychodzące",
        "Waluta w kwocie przychodzącej", "Komentarz",
    ])
    ws_transfers.append([
        date(2026, 1, 25), "Alior", 200.0, "PLN", "Savings", None, "Transfer",
    ])

    path = tmp_path / "test.xlsx"
    wb.save(path)
    return str(path)
